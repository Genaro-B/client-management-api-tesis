from io import BytesIO

from fastapi import APIRouter, Depends, HTTPException, Response, status
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session

from src.schemas.client import CreateClient, UpdateClient, ClientResponse, ProductoAsignado
from typing import List
from src.repositories.client_repo import ClientRepository
from src.repositories.product_repo import ProductRepository
from src.services.client_service import ClientService, EmailAlreadyExists
from src.database.session import get_db
from src.core.exceptions import to_http_exception

router = APIRouter()


@router.get("/inactive")
def list_inactive_clients(
    q: str = None,
    limit: int = 50,
    offset: int = 0,
    db: Session = Depends(get_db),
):
    """Listar clientes inactivos (soft-deleted). Solo acceso admin."""
    repo = ClientRepository(db)
    items, total = repo.list_inactive(limit=limit, offset=offset, nombre=q)
    return {"items": items, "total": total}


@router.patch("/{client_id}/restore", response_model=ClientResponse)
def restore_client(client_id: int, db: Session = Depends(get_db)):
    """Restaurar un cliente previamente eliminado (soft delete). Solo acceso admin."""
    repo = ClientRepository(db)
    client = repo.get_by_id(client_id)
    if client is None:
        raise HTTPException(status_code=404, detail="Cliente no encontrado")
    if client.activo:
        raise HTTPException(status_code=400, detail="El cliente ya se encuentra activo")
    return repo.restore(client)


@router.get("/export")
def export_clients(db: Session = Depends(get_db)):
    """Exportar todos los clientes a un archivo Excel (.xlsx)."""
    repo = ClientRepository(db)
    clients = repo.list_all()

    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Clientes"

    # Encabezados
    headers = ["ID", "Nombre", "Apellido", "Email", "Teléfono", "Fecha Registro", "Estado"]
    header_font = Font(name="Inter", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Datos
    for i, c in enumerate(clients, 2):
        ws.cell(row=i, column=1, value=c.id)
        ws.cell(row=i, column=2, value=c.nombre)
        ws.cell(row=i, column=3, value=c.apellido)
        ws.cell(row=i, column=4, value=c.email)
        ws.cell(row=i, column=5, value=c.telefono or "")
        ws.cell(row=i, column=6, value=c.fecha_registro.strftime("%d/%m/%Y") if c.fecha_registro else "")
        ws.cell(row=i, column=7, value="Activo" if c.activo else "Inactivo")

    # Ajustar ancho de columnas
    col_widths = [6, 25, 25, 35, 18, 18, 12]
    for col, width in enumerate(col_widths, 1):
        ws.column_dimensions[chr(64 + col)].width = width

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=clientes.xlsx"},
    )


@router.post("/", response_model=ClientResponse, status_code=201)
def create_client(payload: CreateClient, response: Response, db: Session = Depends(get_db)):
    """Crear un nuevo cliente. Si el email ya existe, devuelve el cliente existente (200)."""
    repo = ClientRepository(db)
    existing = repo.get_by_email(payload.email)
    if existing:
        response.status_code = status.HTTP_200_OK
        return existing

    service = ClientService(db)
    client = service.create(
        nombre=payload.nombre,
        apellido=payload.apellido,
        telefono=payload.telefono,
        email=payload.email,
    )
    return client


@router.get("/by-email/{email}", response_model=ClientResponse)
def get_client_by_email(email: str, db: Session = Depends(get_db)):
    """Buscar cliente por email exacto. 404 si no existe."""
    repo = ClientRepository(db)
    client = repo.get_by_email(email)
    if client is None:
        raise HTTPException(status_code=404, detail="Cliente no encontrado")
    return client


@router.get("/{client_id}", response_model=ClientResponse)
def get_client(client_id: int, db: Session = Depends(get_db)):
    repo = ClientRepository(db)
    client = repo.get_by_id(client_id)
    if client is None or not client.activo:
        raise HTTPException(status_code=404, detail="Cliente no encontrado")
    return client


@router.get("/")
def list_clients(q: str = None, limit: int = 50, offset: int = 0, db: Session = Depends(get_db)):
    """Listar clientes activos. Soporta búsqueda por nombre y paginación."""
    repo = ClientRepository(db)
    items, total = repo.list(limit=limit, offset=offset, nombre=q)
    return {"items": items, "total": total}


@router.patch("/{client_id}", response_model=ClientResponse)
def update_client(client_id: int, payload: UpdateClient, db: Session = Depends(get_db)):
    service = ClientService(db)
    try:
        result = service.update(client_id, **payload.model_dump(exclude_unset=True))
        if result is None:
            raise HTTPException(status_code=404, detail="Cliente no encontrado")
        return result
    except EmailAlreadyExists as e:
        raise to_http_exception(e)


@router.delete("/{client_id}", status_code=204)
def delete_client(client_id: int, db: Session = Depends(get_db)):
    repo = ClientRepository(db)
    client = repo.get_by_id(client_id)
    if not client:
        raise HTTPException(status_code=404, detail="Cliente no encontrado")
    repo.soft_delete(client)
    return None


# ---------------------------------------------------------------------------
# Productos Asignados
# ---------------------------------------------------------------------------


@router.get("/{client_id}/productos", response_model=List[ProductoAsignado])
def get_client_productos(client_id: int, db: Session = Depends(get_db)):
    """Obtener la lista de productos asignados a un cliente."""
    service = ClientService(db)
    productos = service.get_productos(client_id)
    if productos is None:
        raise HTTPException(status_code=404, detail="Cliente no encontrado")
    return productos


@router.post("/{client_id}/productos", response_model=List[ProductoAsignado], status_code=201)
def add_client_producto(client_id: int, payload: ProductoAsignado, db: Session = Depends(get_db)):
    """Agregar un producto al cliente (o incrementar cantidad si ya existe).
    Valida stock disponible y lo descuenta automáticamente."""
    service = ClientService(db)
    client = service.repo.get_by_id(client_id)
    if not client:
        raise HTTPException(status_code=404, detail="Cliente no encontrado")

    # Check product exists and has enough stock
    product_repo = ProductRepository(db)
    product = product_repo.get_by_id(payload.producto_id)
    if not product:
        raise HTTPException(status_code=400, detail="Producto no encontrado")

    # Calcular cantidad total necesaria: si ya está en el carrito, sumamos
    current = service.repo.get_productos(client)
    existing = next((p for p in current if p["producto_id"] == payload.producto_id), None)
    cantidad_actual = existing["cantidad"] if existing else 0
    cantidad_final = cantidad_actual + payload.cantidad

    if product.stock < payload.cantidad:
        raise HTTPException(
            status_code=400,
            detail=f"Stock insuficiente. Disponible: {product.stock}, solicitado: {payload.cantidad}"
        )

    # Decrement stock
    product.stock -= payload.cantidad
    db.commit()

    # Add to client's list
    result = service.repo.add_producto(client, payload.model_dump())
    return result


@router.put("/{client_id}/productos", response_model=List[ProductoAsignado])
def set_client_productos(client_id: int, payload: List[ProductoAsignado], db: Session = Depends(get_db)):
    """Reemplazar toda la lista de productos asignados (para sync)."""
    service = ClientService(db)
    productos = service.set_productos(client_id, [p.model_dump() for p in payload])
    if productos is None:
        raise HTTPException(status_code=404, detail="Cliente no encontrado")
    return productos


@router.delete("/{client_id}/productos/{producto_id}", status_code=204)
def remove_client_producto(client_id: int, producto_id: int, db: Session = Depends(get_db)):
    """Eliminar un producto específico de la lista del cliente.
    Restaura el stock automáticamente."""
    service = ClientService(db)
    productos = service.get_productos(client_id)
    if productos is None:
        raise HTTPException(status_code=404, detail="Cliente no encontrado")

    # Find the product to get its quantity for stock restoration
    existing = next((p for p in productos if p["producto_id"] == producto_id), None)
    if not existing:
        raise HTTPException(
            status_code=404, detail="Producto no encontrado en la lista del cliente"
        )

    # Restore stock
    product_repo = ProductRepository(db)
    product = product_repo.get_by_id(producto_id)
    if product:
        product.stock += existing["cantidad"]
        db.commit()

    service.remove_producto(client_id, producto_id)
    return None
