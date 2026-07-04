from io import BytesIO

from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session

from src.schemas.product import CreateProduct, UpdateProduct, ProductResponse
from src.repositories.product_repo import ProductRepository
from src.services.product_service import ProductService
from src.database.session import get_db

router = APIRouter()


@router.get("/inactive")
def list_inactive_products(
    q: str = None,
    limit: int = 50,
    offset: int = 0,
    db: Session = Depends(get_db),
):
    """Listar productos inactivos (soft-deleted)."""
    repo = ProductRepository(db)
    items, total = repo.list_inactive(limit=limit, offset=offset, q=q)
    return {"items": items, "total": total}


@router.patch("/{product_id}/restore", response_model=ProductResponse)
def restore_product(product_id: int, db: Session = Depends(get_db)):
    """Restaurar un producto previamente eliminado (soft delete)."""
    repo = ProductRepository(db)
    product = repo.get_by_id(product_id)
    if product is None:
        raise HTTPException(status_code=404, detail="Producto no encontrado")
    if product.activo:
        raise HTTPException(status_code=400, detail="El producto ya se encuentra activo")
    return repo.restore(product)


@router.get("/export")
def export_products(db: Session = Depends(get_db)):
    """Exportar todos los productos activos a un archivo Excel (.xlsx)."""
    repo = ProductRepository(db)
    products = repo.list_all()

    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Productos"

    # Encabezados
    headers = ["ID", "Nombre", "Descripción", "Precio", "Stock", "Categoría", "Fecha Registro", "Estado"]
    header_font = Font(name="Inter", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Datos
    for i, p in enumerate(products, 2):
        ws.cell(row=i, column=1, value=p.id)
        ws.cell(row=i, column=2, value=p.nombre)
        ws.cell(row=i, column=3, value=p.descripcion or "")
        ws.cell(row=i, column=4, value=p.precio)
        ws.cell(row=i, column=5, value=p.stock)
        ws.cell(row=i, column=6, value=p.categoria or "")
        ws.cell(row=i, column=7, value=p.fecha_registro.strftime("%d/%m/%Y") if p.fecha_registro else "")
        ws.cell(row=i, column=8, value="Activo" if p.activo else "Inactivo")

    # Ajustar ancho de columnas
    col_widths = [6, 30, 40, 12, 8, 20, 18, 12]
    for col, width in enumerate(col_widths, 1):
        ws.column_dimensions[chr(64 + col)].width = width

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=productos.xlsx"},
    )


@router.post("/", response_model=ProductResponse, status_code=201)
def create_product(payload: CreateProduct, db: Session = Depends(get_db)):
    """Crear un nuevo producto."""
    service = ProductService(db)
    product = service.create(
        nombre=payload.nombre,
        descripcion=payload.descripcion,
        precio=payload.precio,
        stock=payload.stock,
        categoria=payload.categoria,
    )
    return product


@router.get("/{product_id}", response_model=ProductResponse)
def get_product(product_id: int, db: Session = Depends(get_db)):
    repo = ProductRepository(db)
    product = repo.get_by_id(product_id)
    if product is None or not product.activo:
        raise HTTPException(status_code=404, detail="Producto no encontrado")
    return product


@router.get("/")
def list_products(q: str = None, limit: int = 50, offset: int = 0, db: Session = Depends(get_db)):
    """Listar productos activos. Soporta búsqueda por nombre/categoría y paginación."""
    repo = ProductRepository(db)
    items, total = repo.list(limit=limit, offset=offset, q=q)
    return {"items": items, "total": total}


@router.patch("/{product_id}", response_model=ProductResponse)
def update_product(product_id: int, payload: UpdateProduct, db: Session = Depends(get_db)):
    service = ProductService(db)
    result = service.update(product_id, **payload.model_dump(exclude_unset=True))
    if result is None:
        raise HTTPException(status_code=404, detail="Producto no encontrado")
    return result


@router.delete("/{product_id}", status_code=204)
def delete_product(product_id: int, db: Session = Depends(get_db)):
    repo = ProductRepository(db)
    product = repo.get_by_id(product_id)
    if not product:
        raise HTTPException(status_code=404, detail="Producto no encontrado")
    repo.soft_delete(product)
    return None
